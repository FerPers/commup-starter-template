"""
Genera un Excel de tags realista para un proyecto de Completion & Commissioning
estilo planta GLP / Oil & Gas. ~200 tags distribuidos en 5 disciplinas, varias
áreas, sistemas y subsistemas. Output: demo_tags_glp.xlsx en este mismo dir.
"""
import random
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

random.seed(42)  # determinístico

OUT = Path(__file__).parent / "demo_tags_glp.xlsx"
PROJECT = "Planta GLP - Fase 2"

# ── Hierarchy ────────────────────────────────────────────────────
AREAS = [
    ("A100", "Recepción de Crudo"),
    ("A200", "Tratamiento y Estabilización"),
    ("A300", "Fraccionamiento GLP"),
    ("U100", "Utilidades — Aire/N2/Agua"),
    ("U200", "Utilidades — Generación Eléctrica"),
    ("OS01", "Offsite — Almacenamiento"),
]

# (area_code, system_code, system_name)
SYSTEMS = [
    ("A100", "A100-S01", "Slug Catcher"),
    ("A100", "A100-S02", "Manifold de Recepción"),
    ("A200", "A200-S01", "Separador Trifásico"),
    ("A200", "A200-S02", "Estabilizadora de Crudo"),
    ("A200", "A200-S03", "Tratamiento Amina"),
    ("A300", "A300-S01", "Deetanizadora"),
    ("A300", "A300-S02", "Depropanizadora"),
    ("A300", "A300-S03", "Debutanizadora"),
    ("U100", "U100-S01", "Aire de Instrumentos"),
    ("U100", "U100-S02", "Generación de Nitrógeno"),
    ("U200", "U200-S01", "Turbogenerador"),
    ("U200", "U200-S02", "Distribución 13.8kV"),
    ("OS01", "OS01-S01", "Esferas de GLP"),
    ("OS01", "OS01-S02", "Bombeo a Despacho"),
]

# Disciplinas con sus tag prefixes y tipos
DISC_DEFS = {
    "INST": {
        "name": "Instruments",
        "prefixes": [
            ("FT",   "Transmisor de Flujo"),
            ("PT",   "Transmisor de Presión"),
            ("TT",   "Transmisor de Temperatura"),
            ("LT",   "Transmisor de Nivel"),
            ("FV",   "Válvula de Control de Flujo"),
            ("PV",   "Válvula de Control de Presión"),
            ("LV",   "Válvula de Control de Nivel"),
            ("ESDV", "Válvula de Cierre de Emergencia"),
            ("XV",   "Válvula On/Off"),
            ("PSV",  "Válvula de Alivio"),
        ],
        "manufacturers": [
            ("Emerson",      "Rosemount 3051"),
            ("Yokogawa",     "EJX110A"),
            ("ABB",          "266HSH"),
            ("Endress+Hauser","Cerabar PMC51"),
            ("Honeywell",    "STT850"),
            ("Fisher",       "ED-Series"),
            ("Masoneilan",   "21000"),
        ],
        "fluids": ["Hidrocarburo Líquido", "Gas Natural", "GLP", "Vapor", "Agua de Servicio", "Aire de Instrumentos", "Nitrógeno"],
        "qty": 80,
        "preserve": True,
    },
    "ELEC": {
        "name": "Electrical",
        "prefixes": [
            ("M",    "Motor Eléctrico"),
            ("T",    "Transformador"),
            ("MCC",  "Motor Control Center"),
            ("SWG",  "Switchgear"),
            ("PNL",  "Tablero Eléctrico"),
            ("BAT",  "Banco de Baterías UPS"),
            ("GEN",  "Generador"),
        ],
        "manufacturers": [
            ("Siemens",    "Simotics 1LE1"),
            ("ABB",        "M3BP"),
            ("WEG",        "W22"),
            ("Schneider",  "Altivar"),
            ("Eaton",      "Cutler-Hammer"),
        ],
        "fluids": [],
        "qty": 40,
        "preserve": True,
    },
    "MECH": {
        "name": "Mechanical",
        "prefixes": [
            ("P",  "Bomba Centrífuga"),
            ("C",  "Compresor"),
            ("V",  "Vasija a Presión"),
            ("HX", "Intercambiador de Calor"),
            ("TK", "Tanque Atmosférico"),
            ("F",  "Filtro"),
            ("E",  "Eyector"),
        ],
        "manufacturers": [
            ("Sulzer",      "OHH"),
            ("Flowserve",   "Durco Mark 3"),
            ("KSB",         "RPH"),
            ("Atlas Copco", "ZH"),
            ("Howden",      "Roots"),
            ("Alfa Laval",  "M-Series"),
        ],
        "fluids": ["Hidrocarburo Líquido", "GLP", "Agua de Refrigeración", "Aceite Lubricante"],
        "qty": 35,
        "preserve": True,
    },
    "PIPE": {
        "name": "Piping",
        "prefixes": [
            ("PL",  "Línea de Proceso"),
            ("VL",  "Válvula Manual"),
            ("SP",  "Spool de Tubería"),
            ("STR", "Strainer"),
        ],
        "manufacturers": [
            ("Velan",     "Gate"),
            ("Cameron",   "Ball Valve"),
            ("KITZ",      "Globe"),
            ("Crane",     "Check"),
        ],
        "fluids": ["Hidrocarburo Líquido", "Gas Natural", "GLP", "Vapor"],
        "qty": 30,
        "preserve": False,
    },
    "SAFE": {
        "name": "Safety / Fire",
        "prefixes": [
            ("FP",  "Bomba de Fuego"),
            ("GD",  "Detector de Gas"),
            ("FD",  "Detector de Llama"),
            ("MAC", "Manual Alarm Call"),
            ("DS",  "Rociador / Diluvio"),
            ("FH",  "Hidrante"),
        ],
        "manufacturers": [
            ("Det-Tronics",  "PointWatch IR"),
            ("MSA",          "Ultima X"),
            ("Honeywell",    "Searchpoint"),
            ("Tyco",         "AquaMist"),
            ("Viking",       "Deluge"),
        ],
        "fluids": ["Agua Contra Incendio", "Gas Combustible", "Aire Ambiente"],
        "qty": 15,
        "preserve": True,
    },
}

MOUNTING = ["INLINE", "REMOTE PANEL", "FIELD MOUNTED", "SKID MOUNTED", "WALL MOUNTED", "RACK"]


def gen_tag(prefix: str, idx: int) -> str:
    """Genera un tag plausible: FT-100, ESDV-7621001, M-200A."""
    if prefix in ("ESDV", "PSV"):
        return f"{prefix}-{7621000 + idx}"
    if prefix in ("M", "P"):
        return f"{prefix}-{100 + idx}{random.choice(['A', 'B', ''])}"
    return f"{prefix}-{100 + idx}"


def gen_pid(area: str, idx: int) -> str:
    """P&ID reference: PID-A100-001."""
    return f"PID-{area}-{idx:03d}"


def gen_serial(brand: str) -> str:
    """Serial number plausible."""
    return f"{brand[:3].upper()}-{random.randint(100000, 999999)}"


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Tags"

    # Header de proyecto (rows que el ImportWizard debería ignorar gracias al smart header detection)
    ws["A1"] = f"Proyecto: {PROJECT}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Tag Register — Instrument Index"
    ws["A2"].font = Font(italic=True, color="666666")
    ws["A3"] = "Generado para validación end-to-end del importer (data sintética)"
    ws["A3"].font = Font(italic=True, color="999999")
    # Row 4 vacía intencionalmente para probar el detector de header

    # Header row (row 5)
    headers = [
        "TAG", "DESCRIPCIÓN", "DISCIPLINA",
        "AREA", "NOMBRE ÁREA",
        "SISTEMA", "SUBSISTEMA",
        "FABRICANTE", "MODELO", "SERIE",
        "PRESERVATION", "P&ID", "FLUIDO", "TIPICO MONTAJE",
    ]
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Generate tag rows
    row = 6
    tag_counter_per_prefix: dict[str, int] = {}

    rows_to_emit = []
    for disc_code, disc in DISC_DEFS.items():
        for _ in range(disc["qty"]):
            prefix, base_desc = random.choice(disc["prefixes"])
            tag_counter_per_prefix.setdefault(prefix, 0)
            tag_counter_per_prefix[prefix] += 1
            idx = tag_counter_per_prefix[prefix]

            tag = gen_tag(prefix, idx)
            area_code, area_name = random.choice(AREAS)
            system_candidates = [s for s in SYSTEMS if s[0] == area_code]
            _, sys_code, sys_name = random.choice(system_candidates)
            subsys_code = f"{sys_code}-SS{random.randint(1, 3):02d}"

            mfr, mdl = random.choice(disc["manufacturers"]) if disc["manufacturers"] else ("", "")
            serial = gen_serial(mfr) if mfr else ""

            preservation = "SI" if disc["preserve"] and random.random() < 0.65 else "NO"
            pid = gen_pid(area_code, random.randint(1, 50))
            fluid = random.choice(disc["fluids"]) if disc["fluids"] else ""
            mounting = random.choice(MOUNTING) if disc_code in ("INST", "ELEC", "SAFE") else ""

            rows_to_emit.append([
                tag, base_desc, disc_code,
                area_code, area_name,
                sys_code, subsys_code,
                mfr, mdl, serial,
                preservation, pid, fluid, mounting,
            ])

    random.shuffle(rows_to_emit)  # mezclar disciplinas en orden realista

    for r in rows_to_emit:
        for col, val in enumerate(r, start=1):
            ws.cell(row=row, column=col, value=val)
        row += 1

    # Column widths
    widths = [16, 38, 11, 8, 28, 12, 16, 18, 22, 18, 14, 18, 22, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)].width = w

    ws.freeze_panes = "A6"

    wb.save(OUT)
    total = len(rows_to_emit)
    print(f"✓ Generado {OUT.name}: {total} tags")
    print(f"  Path: {OUT}")
    by_disc = {}
    for r in rows_to_emit:
        by_disc[r[2]] = by_disc.get(r[2], 0) + 1
    print(f"  Distribución: {by_disc}")
    print(f"  Áreas: {len(AREAS)} · Sistemas: {len(SYSTEMS)}")


if __name__ == "__main__":
    main()
